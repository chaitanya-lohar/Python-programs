import openpyxl as xl
from openpyxl.chart import BarChart,Reference
wb=xl.load_workbook("transactions.xlsx")
sheet=wb["Sheet1"]
# cell1=sheet.cell(1,2)
# print(cell1.value)
# print(sheet.max_row)
def automate_process(filename):
    for row in range(2,sheet.max_row+1):
        cell1=sheet.cell(row,3)
        corrected_price=cell1.value*0.9
        corrected_price_cell=sheet.cell(row,4)
        corrected_price_cell.value=corrected_price
        print(corrected_price_cell.value)

    values = Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4, max_col=4)

    chart=BarChart()
    chart.add_data(values)

    sheet.add_chart(chart,'e2')


    wb.save(filename)

automate_process("transactions.xlsx")