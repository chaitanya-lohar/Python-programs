import openpyxl as xl
from openpyxl.chart import BarChart,Reference
def marksheet_generate(filename):
    wb=xl.load_workbook("marks.xlsx")
    sheet=wb["Sheet1"]
    cell1=sheet.cell(1,1)
    print(cell1.value)

    sum=0
    for row in range(3,sheet.max_row-1):
        cell1=sheet.cell(row,2)
        sum=sum+cell1.value
    print(sum)
    sum_cell=sheet.cell(8,2)
    sum_cell.value=sum
    per=sum/5
    per_cell=sheet.cell(9,2)
    per_cell.value=per
    print(per_cell.value)

    values=Reference(sheet,min_row=3,max_row=7,min_col=2,max_col=2)
    chart=BarChart()
    chart.add_data(values)

    sheet.add_chart(chart,"E1")
    wb.save("marks2.xlsx")

marksheet_generate("marks.xlsx")
